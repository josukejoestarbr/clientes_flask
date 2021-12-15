#python -m pip install --trusted-host files.pythonhosted.org --trusted-host pypi.org --trusted-host pypi.python.org 
from flask import Flask, request, render_template,send_file,session,redirect,json , jsonify # Importa a biblioteca
import sqlite3
import time
import sqlite3
from waitress import serve
import pymysql
import credenciais
#porta 3306
#db1 = pymysql.connect(host='localhost', user='root', password='aablcoa12', db='mysql', charset='utf8mb4',port = 3306)
#import pymysql
#porta 3306
from openpyxl import Workbook, load_workbook
Host=''
User=''
Password=''
Db=''
Charset=''
Port=0
Host,User,Password,Db,Charset,Port=credenciais.credencialbanco()
db1 = pymysql.connect(host=Host, user=User, password=Password, db=Db, charset=Charset,port=Port)

cursor = db1.cursor()
#sql=''' 
#CREATE TABLE IF NOT EXISTS tabela(nome varchar(90),date varchar(30), endereco varchar(120),cpf varchar(90),cpfprop varchar(90),
#contrato varchar(10),aluguel varchar(20),date1  varchar(20),date2  varchar(20)) 

#'''

sql7=''' 
CREATE TABLE IF NOT EXISTS blacklist1(nome varchar(70),telefone varchar (30),data varchar (15),detalhes varchar(900),endereco varchar (200) ) 

'''
#cursor.execute(sql)

cursor.execute(sql7)
db1.commit()
db1.close()
from datetime import date
import json


app = Flask(__name__) # Inicializa a aplicação

import os
app.secret_key = "senha_secreta"
import webbrowser
@app.route('/download', methods=['GET','POST']) # Nova rota
def download():
    return send_file('planilhas.xlsx')
@app.route('/', methods=['GET','POST']) # Nova rota
def blacklist():
    retorno=''
    Host=''
    User=''
    Password=''
    Db=''
    Charser=''
    Port=0
    if request.method=='POST':

        if request.form['deleta']=='gerarexcel':
            cpf=request.form['cpf']
            Host,User,Password,Db,Charser,Port=credenciais.credencialbanco()
            db1 = pymysql.connect(host=Host, user=User, password=Password, db=Db, charset=Charset,port=Port)
            cursor = db1.cursor()
            sql1 = "SELECT * FROM blacklist1 WHERE telefone LIKE %s"
            cursor.execute(sql1,(f"%{cpf}%",))
            retorno = cursor.fetchall()
            print (retorno)
            db1.close()
            wb = Workbook()
            ws = wb.active
            a=2
            b=1
            ws.cell(row=1, column=1, value="Nome")
            ws.cell(row=1, column=2, value="Telefone")
            ws.cell(row=1, column=3, value="Telefone")
            ws.cell(row=1, column=4, value="Endereço")
            ws.cell(row=1, column=5, value="Detalhe")
            for i in retorno:
                print (i[0],i[1],i[2])
                ws.cell(row=a, column=b, value=i[0])
                ws.cell(row=a, column=b+1, value=i[1])
                ws.cell(row=a, column=b+2, value=i[2])
                ws.cell(row=a, column=b+3, value=i[4])
                ws.cell(row=a, column=b+4, value=i[3])
                a+=1
            ws.title = "Planilha"
            wb.save("planilhas.xlsx")
            return send_file("planilhas.xlsx")
        if request.form['deleta']=='apaga':
            print ('entrou na blacklist ')
            cpf=request.form['cpf']
            data=request.form['data']
            print ('dados:',cpf,data)
            Host,User,Password,Db,Charser,Port=credenciais.credencialbanco()
            db1 = pymysql.connect(host=Host, user=User, password=Password, db=Db, charset=Charset,port=Port)
            cursor = db1.cursor()
            cursor.execute('DELETE FROM blacklist1 WHERE telefone=%s AND data=%s',(cpf,data))
            db1.commit()
            db1.close()
            Host,User,Password,Db,Charser,Port=credenciais.credencialbanco()
            db1 = pymysql.connect(host=Host, user=User, password=Password, db=Db, charset=Charset,port=Port)
            cursor = db1.cursor()
            sql1 = "SELECT * FROM blacklist1 WHERE telefone LIKE %s"
            cursor.execute(sql1,(f"%{cpf}%",))
            retorno = cursor.fetchall()
            print (retorno)
            db1.close()
            return jsonify(retorno)
        if request.form['busca']=='sim':
            print ('entrou na blacklist')
            cpf=request.form['cpf']
            nome=request.form['nome']
            data=request.form['data']
            fato=request.form['fato']
            endereco=request.form['endereco']
            Host,User,Password,Db,Charser,Port=credenciais.credencialbanco()
            db1 = pymysql.connect(host=Host, user=User, password=Password, db=Db, charset=Charset,port=Port)
            cursor = db1.cursor()
            sql1= """INSERT INTO blacklist1(telefone,nome,data,endereco,detalhes) VALUES (%s,%s,%s,%s,%s)"""
            cursor.execute(sql1, (cpf,nome,data,endereco,fato))
            db1.commit()
            db1.close()
            return jsonify('dados inseridos com sucesso')
    if request.method=='GET':
        if request.args.get('busca')=='sim':
            cpf=request.args.get('cpf')
            print (cpf)
            Host,User,Password,Db,Charser,Port=credenciais.credencialbanco()
            db1 = pymysql.connect(host=Host, user=User, password=Password, db=Db, charset=Charset,port=Port)
            cursor = db1.cursor()
            #LIKE %s',(f"%{cpf}%",))
            sql1 = "SELECT * FROM blacklist1 WHERE telefone LIKE %s"
            cursor.execute(sql1,(f"%{cpf}%",))
            retorno = cursor.fetchall()
            if not retorno:
                retorno='nada'
            print (retorno)
            db1.close()
            return jsonify(retorno)
    return render_template('blacklist.html')


if __name__ == '__main__':
  #serve(app, host='10.75.243.115', port=5000)
  app.run(debug=True) # Executa a aplicação
  #app.run( host='10.75.243.115', port=5000)
